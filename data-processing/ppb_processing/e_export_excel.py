from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from rich import print


COLUMN_DESCRIPTIONS = {
    "Origin Document": "Source document (e.g. 'PPB 2027') from which the mandate row was extracted.",
    "Part In Document": "High-level part of the origin document (e.g. 'Legislative mandates').",
    "File": "Filename of the parsed JSON the row originates from.",
    "Addendum": "Addendum identifier within the origin document, if any.",
    "Part": "Part identifier within the origin document.",
    "Section": "Section number within the origin document.",
    "Section Title": "Title of the section.",
    "Entity-Long": "Full name of the UN entity responsible.",
    "Entity": "Short code / acronym of the UN entity.",
    "Programme": "Programme number.",
    "Programme Title": "Title of the programme.",
    "Subprogramme": "Subprogramme number, if any.",
    "Subprogramme Title": "Title of the subprogramme, if any.",
    "Description": "Mandate description / subject line.",
    "Link": "URL to the underlying mandate document.",
    "Symbol": "UN document symbol of the mandate (short form).",
    "Full Document Symbol": "Full UN document symbol of the mandate.",
    "Source": "Source category of the mandate (e.g. 'General Assembly resolutions').",
}

UN_BLUE = "FF418FDE"
HEADER_FILL = PatternFill(start_color=UN_BLUE, end_color=UN_BLUE, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
INFO_TITLE_FONT = Font(bold=True, size=16, color=UN_BLUE)
INFO_LABEL_FONT = Font(bold=True, size=11)

WRAP_COLUMNS = {
    "Section Title",
    "Entity-Long",
    "Programme Title",
    "Subprogramme Title",
    "Description",
}
MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 10


def _autosize_columns(ws, df):
    for idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        sample = df[col].dropna().astype(str)
        if len(sample) > 0:
            length = int(sample.str.len().quantile(0.95))
        else:
            length = 0
        width = max(
            MIN_COL_WIDTH, min(MAX_COL_WIDTH, max(len(str(col)) + 4, length + 2))
        )
        ws.column_dimensions[letter].width = width


def _style_header(ws, n_cols):
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 28


def _write_data_sheet(wb, df, sheet_name="Mandates"):
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))

    col_idx_by_name = {c: i + 1 for i, c in enumerate(df.columns)}
    link_col = col_idx_by_name.get("Link")

    for row in df.itertuples(index=False, name=None):
        ws.append([("" if pd.isna(v) else v) for v in row])

    n_rows = ws.max_row
    n_cols = ws.max_column

    _autosize_columns(ws, df)

    wrap_cols = {col_idx_by_name[c] for c in WRAP_COLUMNS if c in col_idx_by_name}
    for r in range(2, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=c in wrap_cols,
            )

    if link_col is not None:
        link_font = Font(color=UN_BLUE, underline="single")
        for r in range(2, n_rows + 1):
            cell = ws.cell(row=r, column=link_col)
            url = cell.value
            if isinstance(url, str) and url.startswith(("http://", "https://")):
                cell.hyperlink = url
                cell.font = link_font

    last_col_letter = get_column_letter(n_cols)
    table_ref = f"A1:{last_col_letter}{n_rows}"
    table = Table(displayName="Mandates", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Apply UN-blue header styling on top of the table style
    _style_header(ws, n_cols)
    ws.freeze_panes = "A2"


def _write_info_sheet(wb, df, year, symbol):
    ws = wb.create_sheet("Info", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    ws["A1"] = f"UN Mandates Export — PPB {year}"
    ws["A1"].font = INFO_TITLE_FONT
    ws.merge_cells("A1:B1")

    meta = [
        ("Year", year),
        ("Origin document symbol", symbol),
        ("Total mandates", len(df)),
        (
            "Unique entities",
            df["Entity"].dropna().nunique() if "Entity" in df.columns else "",
        ),
        (
            "Unique sections",
            df["Section"].dropna().nunique() if "Section" in df.columns else "",
        ),
        ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = INFO_LABEL_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Columns").font = INFO_LABEL_FONT
    row += 1
    header_a = ws.cell(row=row, column=1, value="Column")
    header_b = ws.cell(row=row, column=2, value="Description")
    for c in (header_a, header_b):
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
    row += 1
    for col in df.columns:
        ws.cell(row=row, column=1, value=col).font = Font(bold=True)
        desc_cell = ws.cell(row=row, column=2, value=COLUMN_DESCRIPTIONS.get(col, ""))
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.sheet_view.showGridLines = False


def export_excel(year, symbol):
    in_path = Path(f"../data/processed/ppb{year}/all_mandates.csv")
    out_path = Path(f"../data/processed/ppb{year}/all_mandates_formatted.xlsx")
    print(f"[bold]Exporting[/bold] {in_path} → {out_path}")

    df = pd.read_csv(in_path)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _write_info_sheet(wb, df, year, symbol)
    _write_data_sheet(wb, df, sheet_name="Mandates")

    wb.active = wb.sheetnames.index("Mandates")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(
        f"[green]Wrote[/green] {out_path} ({len(df)} rows, {len(df.columns)} columns)"
    )
